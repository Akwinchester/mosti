import openpyxl

import re


# Получить значения ячеек из файла на основе cell_dict
# {'ячейка': "A1"} -> {'ячейка': 'значение'}
def get_values_cell(file_name, cell_dict, calculate=False):
    wb = openpyxl.load_workbook(file_name, data_only=calculate)
    sheet = wb.active

    result = {}
    for key, cell_ref in cell_dict.items():
        cell = sheet[cell_ref]
        result[key] = cell.value

    return result


# Записать значения в ячейки файла на основе cell_value_dict
# {'A1': 'значение1', 'B2': 'значение2'}
def write_to_cell(file_name, cell_value_dict):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    for cell, value in cell_value_dict.items():
        sheet[cell] = value

    wb.save(file_name)


# Получить все числовые ячейки и их левых соседей из файла
def get_numeric_cells(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    sheet = wb.active

    result = {}

    for row in sheet.iter_rows():
        for i in range(1, len(row)):
            cell = row[i]
            left_cell = row[i - 1]

            if left_cell.value is not None and cell.value is not None:
                if isinstance(cell.value, int) or isinstance(cell.value, float):
                    key = cell.coordinate
                    value = (left_cell.value, cell.value)
                    result[key] = value

    return result


# Заменить ссылки на ячейки в формуле значениями
def substitute_values(formula, values_dict):
    if not isinstance(formula, str):
       print(formula)
    else:
        for cell, value in values_dict.items():
            cell_ref = re.escape(cell)   # escape для спецсимволов
            formula = re.sub(rf'\b{cell_ref}\b', str(value[1]), formula)

        formula = formula.replace('=', '')
        formula = formula.replace(' ', '')
        formula = formula.replace('*', '×')

        return formula


# Получить значения формул из файла путем замены ссылок на ячейки
def get_formula_values(name_formul_cell_number, file_name):
    all_name_value = get_numeric_cells(file_name)

    text_formuls = get_values_cell(file_name, name_formul_cell_number)

    for f_k in text_formuls.keys():
        text_formuls[f_k] = substitute_values(text_formuls[f_k], all_name_value)
    return text_formuls

file_name = 'version_2.xlsx'
text_cell = get_values_cell(file_name, {'значение': 'F22'})

# values_cell_dict = get_numeric_cells(file_name)
# replace_cell = substitute_values(text_cell['значение'], values_cell_dict)
# print(replace_cell)
formuls = get_formula_values({'t_p':'M26'}, file_name)
